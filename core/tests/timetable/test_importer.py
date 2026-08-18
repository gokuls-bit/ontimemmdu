import openpyxl
import io
import json
from django.test import TestCase
from core.services.timetable.parser import ParsedEntry
from core.services.timetable.result import ImportResult
from core.services.timetable.validator import validate_parsed_entries
from core.services.timetable.importer import import_timetable
from core.services.timetable.json_importer import import_timetable_json
from timetable.models import Semester, Section, Room, Teacher, Subject, TimetableEntry


class ImporterTestCase(TestCase):
    def test_missing_subject_error(self):
        """6. Missing subject error."""
        result = ImportResult()
        entry = ParsedEntry(
            raw="Dr. Abhishek Bhattacherjee, 357",
            subject_code=None,
            teacher="Dr. Abhishek Bhattacherjee",
            room="357",
            context={'sheet': '3rd', 'cell': 'A1', 'section': '3CSEA1', 'day': 'MON', 'period': 1}
        )
        val = validate_parsed_entries([entry], result)
        self.assertFalse(val)
        self.assertEqual(result.errors[0]['error_code'], "MISSING_SUBJECT")

    def test_missing_teacher_error(self):
        """7. Missing teacher error."""
        result = ImportResult()
        entry = ParsedEntry(
            raw="BCSE-634, 357",
            subject_code="BCSE-634",
            teacher=None,
            room="357",
            context={'sheet': '3rd', 'cell': 'A1', 'section': '3CSEA1', 'day': 'MON', 'period': 1}
        )
        val = validate_parsed_entries([entry], result)
        self.assertFalse(val)
        self.assertEqual(result.errors[0]['error_code'], "MISSING_TEACHER")

    def test_missing_room_error(self):
        """8. Missing room error."""
        result = ImportResult()
        entry = ParsedEntry(
            raw="BCSE-634 Dr. Someone",
            subject_code="BCSE-634",
            teacher="Dr. Someone",
            room=None,
            context={'sheet': '3rd', 'cell': 'A1', 'section': '3CSEA1', 'day': 'MON', 'period': 1}
        )
        val = validate_parsed_entries([entry], result)
        self.assertFalse(val)
        self.assertEqual(result.errors[0]['error_code'], "MISSING_ROOM")

    def test_duplicate_room_booking_error(self):
        """12. Duplicate room booking error."""
        result = ImportResult()
        e1 = ParsedEntry(
            raw="BCSE-634 Dr. Teacher1, 357",
            subject_code="BCSE-634", teacher="Dr. Teacher1", room="357",
            context={'sheet': '3rd', 'cell': 'A1', 'section': '3CSEA1', 'day': 'MON', 'period': 1}
        )
        e2 = ParsedEntry(
            raw="BCSE-501 Dr. Teacher2, 357",
            subject_code="BCSE-501", teacher="Dr. Teacher2", room="357",
            context={'sheet': '3rd', 'cell': 'B1', 'section': '3CSEB1', 'day': 'MON', 'period': 1}
        )
        val = validate_parsed_entries([e1, e2], result)
        self.assertFalse(val)
        self.assertEqual(result.errors[0]['error_code'], "DUPLICATE_ROOM_BOOKING")

    def test_legitimate_merged_section_no_error(self):
        """13. Legitimate shared room / merged section (F,H,J merge)."""
        result = ImportResult()
        e1 = ParsedEntry(
            raw="BCSE-571 Dr. Mohit (F,H,J merge), 269",
            subject_code="BCSE-571", teacher="Dr. Mohit", room="269", merge_group=["F", "H", "J"],
            context={'sheet': '3rd', 'cell': 'A1', 'section': '3CSEA1', 'day': 'MON', 'period': 1}
        )
        e2 = ParsedEntry(
            raw="BCSE-571 Dr. Mohit (F,H,J merge), 269",
            subject_code="BCSE-571", teacher="Dr. Mohit", room="269", merge_group=["F", "H", "J"],
            context={'sheet': '3rd', 'cell': 'B1', 'section': '3CSEB1', 'day': 'MON', 'period': 1}
        )
        val = validate_parsed_entries([e1, e2], result)
        self.assertTrue(val)
        self.assertEqual(result.error_count, 0)

    def test_excel_workbook_import_and_rollback(self):
        """25 & 26. Excel import and atomic transaction rollback on validation failure."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "3rd Semester"

        # Add headers
        ws.cell(row=1, column=1, value="Day")
        ws.cell(row=1, column=2, value="Section")
        ws.cell(row=1, column=3, value="1")
        ws.cell(row=1, column=4, value="2")

        # Row with invalid entry (missing room)
        ws.cell(row=2, column=1, value="Monday")
        ws.cell(row=2, column=2, value="3CSEA1")
        ws.cell(row=2, column=3, value="BCSE-634 Dr. Teacher1")  # No room -> MISSING_ROOM

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "test_3rd.xlsx"

        initial_count = TimetableEntry.objects.count()
        res = import_timetable(buf)

        self.assertFalse(res.success)
        self.assertGreater(res.error_count, 0)
        # Verify rollback occurred (no partial rows created)
        self.assertEqual(TimetableEntry.objects.count(), initial_count)

    def test_json_import_service(self):
        """29. JSON import service."""
        sample_json_data = {
            "semester": "3rd Semester",
            "timetable_entries": [
                {
                    "day": "Monday",
                    "period": 1,
                    "section": "3CSEA1",
                    "raw": "BCSE-634 Dr. Abhishek Bhattacherjee, 357",
                    "entry_type": "LECTURE",
                    "subject_code": "BCSE-634",
                    "teacher": "Dr. Abhishek Bhattacherjee",
                    "room": "357",
                    "source": {"sheet": "3rd Semester", "cell": "C2"}
                }
            ]
        }
        json_file = io.BytesIO(json.dumps(sample_json_data).encode('utf-8'))
        json_file.name = "valid_sample.json"

        res = import_timetable_json(json_file)
        self.assertTrue(res.success)
        self.assertEqual(res.imported_count, 1)

        # Verify record persisted in Django ORM
        entry = TimetableEntry.objects.filter(subject__code="BCSE-634").first()
        self.assertIsNotNone(entry)
        self.assertEqual(entry.room.room_number, "357")
        self.assertEqual(entry.teacher.first_name, "Dr.")
