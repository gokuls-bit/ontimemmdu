import re
import datetime
import openpyxl
from typing import Any, List, Dict, Tuple, Optional
from django.db import transaction
from .result import ImportResult
from .security import validate_excel_security, sanitize_text
from .parser import parse_timetable_cell, ParsedEntry
from .validator import validate_parsed_entries, normalize_day
from timetable.models import (
    Semester, Section, Group, MergeGroup,
    Subject, Teacher, Room, TimeSlot, TimetableEntry
)


def detect_semester_number(text: str) -> Optional[int]:
    """Extracts semester number (e.g. 3, 4, 5, 7) from string."""
    if not text:
        return None
    match = re.search(r'\b([1-8])(?:st|nd|rd|th)?\s*(?:Sem|Semester)\b', text, re.IGNORECASE)
    if match:
        return int(match.group(1))
    # Fallback search for digit followed by semester
    match_fallback = re.search(r'\b([1-8])\b', text)
    if match_fallback:
        return int(match_fallback.group(1))
    return None


def parse_period_timing(timing_str: str) -> Tuple[datetime.time, datetime.time]:
    """Parses timing strings like '08:40am-09:40am' into start_time, end_time."""
    default_start = datetime.time(9, 0)
    default_end = datetime.time(10, 0)
    if not timing_str:
        return default_start, default_end

    parts = re.split(r'[-–to]+', timing_str.strip())
    if len(parts) == 2:
        try:
            fmt = "%I:%M%p"
            t1_str = parts[0].strip().replace(" ", "").upper()
            t2_str = parts[1].strip().replace(" ", "").upper()
            t1 = datetime.datetime.strptime(t1_str, fmt).time()
            t2 = datetime.datetime.strptime(t2_str, fmt).time()
            return t1, t2
        except Exception:
            pass

    return default_start, default_end


def import_timetable(file_input: Any, academic_year: str = "2026-27") -> ImportResult:
    """
    Imports an Excel timetable workbook into the Django/PostgreSQL database.
    
    Order of operations:
    1. Security validation
    2. Open workbook safely
    3. Discover sheets & semester
    4. Discover period/day structure & sections
    5. Parse all relevant cells
    6. Validate all entries & room conflicts
    7. Wrap database persistence in transaction.atomic()
    """
    filename = getattr(file_input, 'name', str(file_input))
    result = ImportResult(file_name=filename)

    # 1. Security Check
    if not validate_excel_security(file_input, result):
        return result

    # 2. Open Workbook Safely
    try:
        wb = openpyxl.load_workbook(
            file_input,
            read_only=False,
            data_only=True,
            keep_links=False
        )
    except Exception as e:
        result.add_error("CORRUPTED_WORKBOOK", f"Failed to open Excel workbook: {str(e)}")
        return result

    parsed_entries: List[ParsedEntry] = []
    discovered_sections: Set[str] = set()
    discovered_semester_num: Optional[int] = None
    discovered_semester_name: str = ""

    # 3. Inspect Worksheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Determine Semester from Sheet Name or Top Header
        sem_num = detect_semester_number(sheet_name)
        if not sem_num:
            # Check first 5 rows for header text
            for r in range(1, 6):
                for c in range(1, 10):
                    cell_val = str(ws.cell(row=r, column=c).value or '')
                    found_sem = detect_semester_number(cell_val)
                    if found_sem:
                        sem_num = found_sem
                        break
                if sem_num:
                    break

        if sem_num and not discovered_semester_num:
            discovered_semester_num = sem_num
            discovered_semester_name = f"{sem_num}th Semester"

        # Determine Layout Structure (Periods, Days, Sections)
        header_row = None
        period_cols: Dict[int, int] = {}  # col_idx -> period_num
        section_col = None
        day_col = None

        # Standard classwise sheet layout search
        for r in range(1, 12):
            row_vals = [str(ws.cell(row=r, column=c).value or '').strip() for c in range(1, 25)]
            # Search for period markers (1, 2, 3, 4, 5, 6, 7, 8)
            p_indices = {}
            for col_idx, val in enumerate(row_vals, start=1):
                clean_v = val.upper()
                if clean_v.isdigit() and 1 <= int(clean_v) <= 12:
                    p_indices[col_idx] = int(clean_v)
                elif "PERIOD" in clean_v:
                    m = re.search(r'\b([1-9]|1[0-2])\b', clean_v)
                    if m:
                        p_indices[col_idx] = int(m.group(1))

            if len(p_indices) >= 3:
                header_row = r
                period_cols = p_indices
                break

        if not header_row or not period_cols:
            # Sheet does not contain a standard period header, check column-wise sections
            continue

        # Find Day and Section columns
        for c in range(1, 6):
            col_vals = [str(ws.cell(row=r, column=c).value or '').upper() for r in range(1, 15)]
            if any("DAY" in v for v in col_vals):
                day_col = c
            if any("SEC" in v or "SECTION" in v or "CLASS" in v for v in col_vals):
                section_col = c

        if not day_col:
            day_col = 1
        if not section_col:
            section_col = 2 if day_col == 1 else 1

        # Iterate Rows to Extract Timetable Cells
        current_day = "MONDAY"
        for r in range(header_row + 1, ws.max_row + 1):
            day_val = str(ws.cell(row=r, column=day_col).value or '').strip()
            if day_val and normalize_day(day_val) in {'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN'}:
                current_day = day_val

            sec_val = sanitize_text(ws.cell(row=r, column=section_col).value)
            if not sec_val or sec_val.upper() in {"DAY", "SECTION", "PERIOD", "TIME", "TIMING"}:
                continue

            discovered_sections.add(sec_val)

            # Parse each period column in this row
            for c_idx, p_num in period_cols.items():
                cell_obj = ws.cell(row=r, column=c_idx)
                raw_cell_val = cell_obj.value
                coord = cell_obj.coordinate

                ctx = {
                    'sheet': sheet_name,
                    'cell': coord,
                    'row': r,
                    'column': c_idx,
                    'section': sec_val,
                    'day': current_day,
                    'period': p_num,
                }

                parsed = parse_timetable_cell(raw_cell_val, context=ctx)
                parsed_entries.append(parsed)

    result.parsed_count = len(parsed_entries)
    result.sections_count = len(discovered_sections)
    result.semester = discovered_semester_name or (f"{discovered_semester_num}th Semester" if discovered_semester_num else "Unknown Semester")

    # Check if semester was determined
    if not discovered_semester_num:
        result.add_error("INVALID_SEMESTER", "Could not reliably determine semester from workbook headers or sheets.")
        return result

    # 4. Run Validation Pipeline
    is_valid = validate_parsed_entries(parsed_entries, result)
    if not is_valid or result.error_count > 0:
        return result

    # 5. Atomic Database Persistence
    try:
        with transaction.atomic():
            sem_obj, _ = Semester.objects.get_or_create(
                number=discovered_semester_num,
                academic_year=academic_year,
                defaults={'is_active': True}
            )

            # Create default time slots if missing
            slot_map: Dict[Tuple[str, int], TimeSlot] = {}
            for d_code in ['MON', 'TUE', 'WED', 'THU', 'FRI']:
                for p in range(1, 9):
                    s_hr = 8 + p if p <= 4 else 9 + p
                    ts, _ = TimeSlot.objects.get_or_create(
                        day=d_code,
                        period=p,
                        defaults={
                            'start_time': datetime.time(s_hr, 0),
                            'end_time': datetime.time(s_hr + 1, 0)
                        }
                    )
                    slot_map[(d_code, p)] = ts

            section_objs: Dict[str, Section] = {}
            for sec_name in discovered_sections:
                sec_obj, _ = Section.objects.get_or_create(
                    name=sec_name,
                    semester=sem_obj,
                    defaults={'capacity': 60}
                )
                section_objs[sec_name] = sec_obj

            merge_group_objs: Dict[str, MergeGroup] = {}
            entries_to_create = []

            for entry in parsed_entries:
                if entry.is_break:
                    continue

                ctx = entry.context
                sec_name = ctx.get('section')
                day_code = normalize_day(ctx.get('day'))
                period_num = ctx.get('period')

                sec_obj = section_objs.get(sec_name)
                ts_obj = slot_map.get((day_code, period_num)) or TimeSlot.objects.filter(day=day_code, period=period_num).first()

                if not ts_obj:
                    ts_obj = TimeSlot.objects.create(
                        day=day_code,
                        period=period_num,
                        start_time=datetime.time(9, 0),
                        end_time=datetime.time(10, 0)
                    )

                # Subject
                sub_code = entry.subject_code or "SPEC-00"
                sub_obj, _ = Subject.objects.get_or_create(
                    code=sub_code,
                    defaults={
                        'name': entry.subject_name or sub_code,
                        'short_name': sub_code,
                        'subject_type': entry.class_type if entry.class_type in Subject.SubjectType.values else Subject.SubjectType.THEORY,
                        'semester': sem_obj
                    }
                )

                # Teacher
                t_name = entry.teacher or "Faculty Staff"
                fname = t_name.split()[0] if t_name else "Faculty"
                lname = " ".join(t_name.split()[1:]) if len(t_name.split()) > 1 else "Staff"
                emp_id = f"EMP-{hash(t_name) % 100000:05d}"

                teacher_obj = Teacher.objects.filter(first_name=fname, last_name=lname).first()
                if not teacher_obj:
                    teacher_obj, _ = Teacher.objects.get_or_create(
                        employee_id=emp_id,
                        defaults={
                            'first_name': fname,
                            'last_name': lname,
                            'email': f"{fname.lower()}.{lname.lower().replace(' ', '')}@cse.edu",
                            'designation': 'Assistant Professor'
                        }
                    )

                # Room
                r_num = entry.room or "C-301"
                room_obj, _ = Room.objects.get_or_create(
                    room_number=r_num,
                    defaults={
                        'building': 'Engineering Block C',
                        'room_type': Room.RoomType.LAB if entry.class_type == 'LAB' else Room.RoomType.LECTURE_HALL,
                        'capacity': 35 if entry.class_type == 'LAB' else 70
                    }
                )

                # Handle Merged Group
                merge_group_obj = None
                if entry.merge_group:
                    mg_name_str = "_".join(sorted(entry.merge_group))
                    mg_key = f"{sem_obj.number}Sem-{mg_name_str}"
                    if mg_key not in merge_group_objs:
                        mg_obj, _ = MergeGroup.objects.get_or_create(
                            name=mg_key,
                            defaults={'description': f"Merged group for sections/groups: {', '.join(entry.merge_group)}"}
                        )
                        merge_group_objs[mg_key] = mg_obj

                        # Link component groups to merge group
                        if sec_obj:
                            for grp_letter in entry.merge_group:
                                grp_obj, _ = Group.objects.get_or_create(name=grp_letter, section=sec_obj)
                                mg_obj.groups.add(grp_obj)

                    merge_group_obj = merge_group_objs[mg_key]
                    result.merge_groups_count = len(merge_group_objs)

                tt_entry = TimetableEntry(
                    semester=sem_obj,
                    section=sec_obj if not merge_group_obj else sec_obj,
                    merge_group=merge_group_obj,
                    subject=sub_obj,
                    teacher=teacher_obj,
                    room=room_obj,
                    time_slot=ts_obj,
                    day=day_code,
                    period=period_num,
                    start_time=ts_obj.start_time,
                    end_time=ts_obj.end_time,
                    class_type=TimetableEntry.ClassType.LAB if entry.class_type == 'LAB' else TimetableEntry.ClassType.LECTURE
                )
                entries_to_create.append(tt_entry)

            # Save all timetable entries
            for tt in entries_to_create:
                tt.save()

            result.imported_count = len(entries_to_create)
            result.success = True

    except Exception as e:
        result.add_error("TRANSACTION_FAILED", f"Database import transaction failed: {str(e)}")
        result.success = False

    return result
